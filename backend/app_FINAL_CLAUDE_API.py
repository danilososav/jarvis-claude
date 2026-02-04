"""
JARVIS - Backend Flask con PostgreSQL + Autenticación + Fuzzy Matching + Tablas Dinámicas
Sistema BI conversacional para agencia de medios (Paraguay)
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from dotenv import load_dotenv
import logging
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
from functools import wraps
from fuzzywuzzy import fuzz
import pandas as pd
import io
import json
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage
from reportlab.lib.units import inch
from io import BytesIO
import base64
from datetime import datetime
import plotly.graph_objects as go
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart



# SQLAlchemy
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime
from sqlalchemy.orm import declarative_base, sessionmaker
from sqlalchemy import text
from app_claude_integration import handle_query_with_claude

import base64
from io import BytesIO
import plotly.graph_objects as go
        


load_dotenv()

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Flask app
app = Flask(__name__)
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# ==================== DATABASE ====================

DB_USER = os.getenv('PG_USER', 'postgres')
DB_PASS = os.getenv('PG_PASS', '12345')
DB_HOST = os.getenv('PG_HOST', 'localhost')
DB_PORT = os.getenv('PG_PORT', '5432')
DB_NAME = os.getenv('PG_DB', 'jarvis')

DB_URL = f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(DB_URL, pool_pre_ping=True)

# SQLAlchemy models
Base = declarative_base()

class User(Base):
    __tablename__ = 'users'
    id = Column(Integer, primary_key=True)
    username = Column(String(50), unique=True)
    password_hash = Column(String(255))
    role = Column(String(20), default='normal')
    created_at = Column(DateTime, default=datetime.now)

class Conversation(Base):
    __tablename__ = 'conversations'
    id = Column(Integer, primary_key=True)
    user_id = Column(Integer)
    session_id = Column(String(50))  # NUEVO: agrupar chats
    query = Column(Text)
    response = Column(Text)
    query_type = Column(String(50))
    chart_config = Column(Text)
    chart_data = Column(Text)
    created_at = Column(DateTime, default=datetime.now)

class TrainerFeedback(Base):
    __tablename__ = 'trainer_feedback'
    id = Column(Integer, primary_key=True)
    conversation_id = Column(Integer)
    user_id = Column(Integer)
    original_query = Column(Text)
    original_response = Column(Text)
    corrected_response = Column(Text)
    feedback_type = Column(String(50))
    notes = Column(Text)
    created_at = Column(DateTime, default=datetime.now)

class DynamicTable(Base):
    __tablename__ = 'dynamic_tables'
    id = Column(Integer, primary_key=True)
    table_name = Column(String(100), unique=True)
    columns = Column(Text)
    created_by = Column(Integer)
    created_at = Column(DateTime, default=datetime.now)

class AuditLog(Base):
    __tablename__ = 'audit_logs'
    id = Column(Integer, primary_key=True)
    user_id = Column(Integer)
    username = Column(String(50))
    action = Column(String(100))
    details = Column(Text)
    ip_address = Column(String(45))
    created_at = Column(DateTime, default=datetime.now)

# Create tables
try:
    Base.metadata.create_all(engine)
    logger.info("✅ Tablas creadas/verificadas")
except Exception as e:
    logger.error(f"Error creando tablas: {e}")

Session = sessionmaker(bind=engine)

# ==================== AUTHENTICATION ====================

SECRET_KEY = os.getenv('SECRET_KEY', 'jarvis-secret-key-2026')

def generate_token(user_id):
    """Genera JWT token"""
    payload = {
        'user_id': user_id,
        'exp': datetime.utcnow() + timedelta(days=7)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

def verify_token(token):
    """Verifica JWT token"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        return payload['user_id']
    except:
        return None

def token_required(f):
    """Decorator para proteger endpoints"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'Token requerido'}), 401
        
        token = token.replace('Bearer ', '')
        user_id = verify_token(token)
        
        if not user_id:
            return jsonify({'error': 'Token inválido'}), 401
        
        return f(user_id, *args, **kwargs)
    return decorated

# ==================== FUZZY MATCHING ====================

def log_audit(user_id, username, action, details='', ip_address=''):
    """Registra auditoría"""
    try:
        session = Session()
        audit = AuditLog(
            user_id=user_id,
            username=username,
            action=action,
            details=details,
            ip_address=ip_address
        )
        session.add(audit)
        session.commit()
        session.close()
    except Exception as e:
        logger.error(f"Error en auditoría: {e}")

def should_generate_chart(user_query):
    """Detecta si el usuario específicamente pide un gráfico"""
    chart_keywords = ['gráfico', 'grafico', 'chart', 'visualiz', 'gráfica', 'grafica', 'mostrar en gráfico']
    return any(keyword in user_query.lower() for keyword in chart_keywords)

def find_similar_feedback(user_query, category=None):
    """Busca feedback similar con threshold dinámico y categoría"""
    try:
        session = Session()
        
        # Si hay categoría, filtrar por esa primero
        if category:
            feedbacks = session.query(TrainerFeedback)\
                .filter(TrainerFeedback.category == category)\
                .all()
        else:
            feedbacks = session.query(TrainerFeedback).all()
        
        session.close()
        
        best_match = None
        best_score = 0
        
        # Threshold dinámico según longitud de query
        query_len = len(user_query.split())
        if query_len <= 3:
            threshold = 75  # Queries cortas: menos exigentes
        elif query_len <= 7:
            threshold = 80  # Queries medianas
        else:
            threshold = 85  # Queries largas: más exigentes
        
        for fb in feedbacks:
            if fb.original_query:
                # Usar token_set_ratio para más flexibilidad
                score = fuzz.token_set_ratio(user_query.lower(), fb.original_query.lower())
                if score > best_score:
                    best_score = score
                    best_match = fb
        
        if best_match and best_score >= threshold:
            logger.info(f"✅ Feedback similar encontrado: score {best_score}, categoría: {best_match.category}")
            best_match.similarity_score = best_score
            return best_match
        
        logger.info(f"❌ No hay feedback similar (mejor score: {best_score}, threshold: {threshold})")
        return None
    except Exception as e:
        logger.error(f"Error en fuzzy matching: {e}")
        return None

# ==================== QUERIES A LA BD ====================

def mock_claude_response(query_type, rows, user_query):
    """Respuestas inteligentes sin usar API Claude"""
    
    if query_type == "corrected":
        return "Respuesta corregida por entrenador"
    
    elif query_type == "dynamic_table":
        if rows:
            return f"Encontré {len(rows)} registros en esa tabla. Primeros registros: {str(rows[:2])}"
        return "No hay datos en esa tabla."
    
    elif query_type == "chart":
        if rows:
            total = sum(r.get("facturacion", 0) for r in rows)
            return f"Aquí te muestro un gráfico con los {len(rows)} clientes principales. Facturación total: {total:,.0f} Gs"
        return "No hay datos disponibles para el gráfico."
    
    elif query_type == "ranking":
        if rows:
            total_facturacion = sum(r.get("facturacion", 0) for r in rows)
            parts = [f"{i+1}. {r['cliente']}: {r['facturacion']:,.0f} Gs ({r.get('market_share', 0):.2f}%)" 
                    for i, r in enumerate(rows)]
            return f"Top {len(rows)} clientes por facturación:\n" + "\n".join(parts) + f"\nTotal: {total_facturacion:,.0f} Gs"
        return "No encontré datos de clientes."
    
    elif query_type == "facturacion":
        if rows:
            r = rows[0]
            return f"**{r['cliente']}** facturó **{r['facturacion']:,.0f} Gs** con un market share de {r.get('market_share', 0):.2f}%. Promedio mensual: {r.get('promedio_mensual', 0):,.0f} Gs."
        return "No tengo datos de facturación para ese cliente."
    
    else:
        return "Consulta procesada."

def get_top_clientes_enriched(query):
    """Top 5 clientes por facturación"""
    try:
        with engine.connect() as conn:
            stmt = text("""
                SELECT 
                    d.nombre_canonico,
                    SUM(f.facturacion)::float as facturacion,
                    COUNT(*) as registros,
                    (SUM(f.facturacion) / NULLIF((SELECT SUM(facturacion) FROM fact_facturacion WHERE facturacion > 0), 0) * 100)::float as market_share
                FROM dim_anunciante d
                LEFT JOIN fact_facturacion f ON d.anunciante_id = f.anunciante_id
                WHERE f.facturacion > 0
                GROUP BY d.anunciante_id, d.nombre_canonico
                HAVING SUM(f.facturacion) > 0
                ORDER BY facturacion DESC
                LIMIT 5
            """)
            rows = conn.execute(stmt).fetchall()
            
            result = []
            for r in rows:
                result.append({
                    "cliente": r[0] or "Sin nombre",
                    "facturacion": float(r[1]) if r[1] else 0,
                    "registros": int(r[2]) if r[2] else 0,
                    "market_share": float(r[3]) if r[3] else 0
                })
            return result
    except Exception as e:
        logger.error(f"Error Top Clientes: {e}")
        return []

def get_facturacion_enriched(query):
    """Facturación de un cliente específico"""
    query_limpio = query.replace('?', '').replace('!', '').replace(',', '')
    palabras = query_limpio.split()
    cliente = " ".join([p for p in palabras if p.isupper()])
    
    if not cliente:
        return []
    
    try:
        with engine.connect() as conn:
            stmt = text("""
                SELECT 
                    d.nombre_canonico,
                    SUM(f.facturacion)::float as facturacion,
                    AVG(f.facturacion)::float as promedio_mensual,
                    (SUM(f.facturacion) / (SELECT SUM(facturacion) FROM fact_facturacion) * 100)::float as market_share
                FROM dim_anunciante d
                LEFT JOIN fact_facturacion f ON d.anunciante_id = f.anunciante_id AND f.facturacion > 0
                WHERE UPPER(d.nombre_canonico) LIKE UPPER(:cliente)
                GROUP BY d.anunciante_id, d.nombre_canonico
            """)
            rows = conn.execute(stmt, {"cliente": f"%{cliente}%"}).fetchall()
            
            result = []
            for r in rows:
                result.append({
                    "cliente": r[0],
                    "facturacion": float(r[1]) if r[1] else 0,
                    "promedio_mensual": float(r[2]) if r[2] else 0,
                    "market_share": float(r[3]) if r[3] else 0
                })
            return result
    except Exception as e:
        logger.error(f"Error Facturación: {e}")
        return []

# ==================== ENDPOINTS ====================

@app.route('/api/health', methods=['GET'])
def health():
    """Health check"""
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return jsonify({"status": "✅ OK", "db": "connected"}), 200
    except:
        return jsonify({"status": "❌ ERROR", "db": "disconnected"}), 500

# ==================== AUTHENTICATION ENDPOINTS ====================

@app.route('/api/auth/register', methods=['POST'])
def register():
    """Registrar nuevo usuario"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return jsonify({'error': 'Username y password requeridos'}), 400
        
        if len(password) < 6:
            return jsonify({'error': 'Password mínimo 6 caracteres'}), 400
        
        session = Session()
        
        existing = session.query(User).filter_by(username=username).first()
        if existing:
            session.close()
            return jsonify({'error': 'Usuario ya existe'}), 409
        
        user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role='normal'
        )
        session.add(user)
        session.commit()
        user_id = user.id
        session.close()
        
        token = generate_token(user_id)
        
        logger.info(f"✅ Usuario registrado: {username}")
        
        return jsonify({
            'success': True,
            'user_id': user_id,
            'username': username,
            'role': 'normal',
            'token': token
        }), 201
        
    except Exception as e:
        logger.error(f"Error registrando usuario: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login de usuario"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return jsonify({'error': 'Username y password requeridos'}), 400
        
        session = Session()
        user = session.query(User).filter_by(username=username).first()
        session.close()
        
        if not user or not check_password_hash(user.password_hash, password):
            return jsonify({'error': 'Credenciales inválidas'}), 401
        
        token = generate_token(user.id)
        
        logger.info(f"✅ Login: {username}")
        
        return jsonify({
            'success': True,
            'user_id': user.id,
            'username': user.username,
            'role': user.role,
            'token': token
        }), 200
        
        log_audit(user.id, user.username, 'LOGIN', ip_address=request.remote_addr)
    except Exception as e:
        logger.error(f"Error en login: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/verify', methods=['GET'])
@token_required
def verify_user(user_id):
    """Verificar token y obtener info del usuario"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        session.close()
        
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        return jsonify({
            'success': True,
            'user_id': user.id,
            'username': user.username,
            'role': user.role
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== QUERY ENDPOINTS ====================
@app.route('/api/chat/history', methods=['GET'])
@token_required
def get_chat_history(user_id):
    """Obtener historial agrupado por sesión"""
    try:
        session = Session()
        
        # Obtener todas las conversaciones
        conversations = session.query(Conversation)\
            .filter_by(user_id=user_id)\
            .order_by(Conversation.created_at.desc())\
            .all()
        
        # Agrupar por session_id
        grouped = {}
        for c in conversations:
            sid = c.session_id or f"session_{user_id}"
            if sid not in grouped:
                grouped[sid] = []
            grouped[sid].append({
                "id": c.id,
                "query": c.query,
                "response": c.response,
                "query_type": c.query_type,
                "chart_config": json.loads(c.chart_config) if c.chart_config else None,
                "rows": json.loads(c.chart_data) if c.chart_data else None,
                "created_at": c.created_at.isoformat()
            })
        
        # Convertir a lista de sesiones
        result = []
        for session_id, messages in grouped.items():
            result.append({
                "session_id": session_id,
                "messages": messages,
                "created_at": messages[0]["created_at"]
            })
        
        session.close()
        return jsonify({"success": True, "sessions": result}), 200
    except Exception as e:
        logger.error(f"Error obteniendo historial: {e}")
        return jsonify({"error": str(e)}), 500
    
@app.route('/api/chat/history/<int:conv_id>', methods=['DELETE'])
@token_required
def delete_conversation(user_id, conv_id):
    """Eliminar una conversación"""
    session = None
    try:
        session = Session()
        
        conversation = session.query(Conversation).filter_by(
            id=conv_id, 
            user_id=user_id
        ).first()
        
        if not conversation:
            session.close()
            return jsonify({'error': 'Conversación no encontrada'}), 404
        
        session.delete(conversation)
        session.commit()
        session.close()
        
        logger.info(f"✅ Conversación {conv_id} eliminada")
        
        return jsonify({'success': True}), 200
        
    except Exception as e:
        logger.error(f"Error eliminando conversación: {e}")
        if session:
            session.rollback()
            session.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/query', methods=['POST'])
@token_required
def query(user_id):
    data = request.json
    user_query = data.get('query', '').strip()
    session_id = data.get('session_id', str(user_id))  # NUEVO
    """Procesar query con detección automática"""
    session = Session()
    user = session.query(User).filter_by(id=user_id).first()
    username = user.username if user else 'unknown'
    session.close()
    
    data = request.json
    user_query = data.get('query', '').strip()
    log_audit(user_id, user.username, 'QUERY', details=user_query[:100], ip_address=request.remote_addr)

    if not user_query:
        return jsonify({"error": "Query vacío"}), 400
    
    try:
        query_lower = user_query.lower()
        rows = []
        query_type = "generico"
        chart_config = None
        response_text = None
        
        # BUSCAR FEEDBACK SIMILAR PRIMERO
        similar_feedback = find_similar_feedback(user_query)
        if similar_feedback:
            response_text = similar_feedback.corrected_response
            query_type = "corrected"
            logger.info(f"✅ Usando respuesta corregida por trainer")
        else:
            # DETECCIÓN DE TABLAS DINÁMICAS
            dynamic_session = Session()
            try:
                dynamic_tables = dynamic_session.query(DynamicTable).all()
                for dt in dynamic_tables:
                    if dt.table_name in query_lower:
                        query_type = "dynamic_table"
                        try:
                            with engine.connect() as conn:
                                stmt = text(f"SELECT * FROM {dt.table_name} LIMIT 10")
                                result = conn.execute(stmt).fetchall()
                                rows = [dict(row._mapping) for row in result]
                                logger.info(f"🔍 Tabla dinámica detectada: {dt.table_name}")
                        except Exception as e:
                            logger.error(f"Error queryando tabla dinámica: {e}")
                            rows = []
                        break
            finally:
                dynamic_session.close()
            
            # Si no encontró tabla dinámica, continuar con otras detecciones
            if query_type == "generico":
                # DETECCIÓN DE GRÁFICOS - SOLO si user explícitamente pide gráfico
                if should_generate_chart(user_query):
                    query_type = "chart"
                    
                    if "barra" in query_lower or "bar" in query_lower:
                        chart_config = {"type": "bar"}
                    elif "pie" in query_lower or "torta" in query_lower or "circular" in query_lower:
                        chart_config = {"type": "pie"}
                    elif "línea" in query_lower or "linea" in query_lower or "line" in query_lower:
                        chart_config = {"type": "line"}
                    else:
                        chart_config = {"type": "bar"}  # Default a bar
                    
                    rows = get_top_clientes_enriched(user_query)
                    logger.info(f"🔍 Detectado: gráfico {chart_config['type']} - rows: {len(rows)}")
                
                # DETECCIÓN DE RANKING - SIN gráfico, solo tabla
                elif any(w in query_lower for w in ["top", "ranking", "principal", "importante", "mayor", "más", "clientes"]):
                    query_type = "ranking"
                    rows = get_top_clientes_enriched(user_query)
                    chart_config = None  # IMPORTANTE: NO generar gráfico
                    logger.info(f"🔍 Detectado: ranking (sin gráfico) - rows: {len(rows)}")
                
                # DETECCIÓN DE FACTURACIÓN
                elif any(w in query_lower for w in ["cuánto", "cuanto", "factur", "how much"]):
                    query_type = "facturacion"
                    rows = get_facturacion_enriched(user_query)
                    logger.info(f"🔍 Detectado: facturacion - rows: {len(rows)}")
                
                else:
                    return jsonify({"success": False, "response": "Consulta no reconocida. Prueba: 'Top 5 clientes', 'Gráfico de barras', 'Cuánto facturó CERVEPAR?'"}), 200
            
            # Generar respuesta si no es corregida
            response_text = mock_claude_response(query_type, rows, user_query)
        
        # Guardar en BD
        # Guardar en BD
        session = None
        try:
            session = Session()
            conversation = Conversation(
            user_id=user_id,
            session_id=session_id,  # NUEVO
            query=user_query,
            response=response_text,
            query_type=query_type,
            chart_config=json.dumps(chart_config) if chart_config else None,
            chart_data=json.dumps(rows) if rows else None
        )
            session.add(conversation)
            session.commit()
            conv_id = conversation.id
            logger.info(f"✅ Conversación guardada: {conv_id}")
        except Exception as db_err:
            logger.error(f"❌ Error guardando en BD: {db_err}")
            if session:
                session.rollback()
            conv_id = None
        finally:
            if session:
                session.close()
        
        return jsonify({
            "success": True,
            "response": response_text,
            "query_type": query_type,
            "chart_config": chart_config,
            "rows": rows,
            "conversation_id": conv_id
        }), 200
        
    except Exception as e:
        logger.error(f"Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# ==================== TRAINER ENDPOINTS ====================
@app.route('/api/trainer/feedback', methods=['POST'])
@token_required
def submit_trainer_feedback(user_id):
    """Enviar corrección/feedback como trainer con categoría y tags"""
    session = None
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        username = user.username if user else 'unknown'
        
        if not user or user.role != 'trainer':
            session.close()
            return jsonify({'error': 'Solo trainers pueden enviar feedback'}), 403
        
        data = request.json
        feedback = TrainerFeedback(
            conversation_id=data.get('conversation_id'),
            user_id=user_id,
            original_query=data.get('original_query'),
            original_response=data.get('original_response'),
            corrected_response=data.get('corrected_response'),
            feedback_type=data.get('feedback_type', 'correction'),
            notes=data.get('notes', ''),
            category=data.get('category'),  # NUEVO
            tags=json.dumps(data.get('tags', [])),  # NUEVO - JSON array
            chart_config=json.dumps(data.get('chart_config')) if data.get('chart_config') else None,  # NUEVO
            query_type=data.get('query_type')  # NUEVO
        )
        session.add(feedback)
        session.commit()
        feedback_id = feedback.id
        session.close()
        
        # Log después de cerrar sesión
        log_audit(user_id, username, 'FEEDBACK', details=f"Feedback ID: {feedback_id}, Categoría: {data.get('category')}", ip_address=request.remote_addr)
        
        logger.info(f"✅ Trainer feedback guardado: {feedback_id}, categoría: {data.get('category')}")
        
        return jsonify({'success': True, 'feedback_id': feedback_id}), 201
        
    except Exception as e:
        logger.error(f"Error en trainer feedback: {e}")
        if session:
            session.rollback()
            session.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/trainer/feedback', methods=['GET'])
@token_required
def get_trainer_feedback(user_id):
    """Obtener feedback guardado con opción de filtrar por categoría"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return jsonify({'error': 'Solo trainers pueden ver feedback'}), 403
        
        # Parámetro opcional para filtrar por categoría
        category = request.args.get('category')
        
        query = session.query(TrainerFeedback).order_by(TrainerFeedback.created_at.desc())
        
        if category:
            query = query.filter_by(category=category)
        
        feedbacks = query.limit(100).all()
        
        result = [{
            "id": f.id,
            "conversation_id": f.conversation_id,
            "original_query": f.original_query,
            "original_response": f.original_response[:100] if f.original_response else None,
            "corrected_response": f.corrected_response[:100] if f.corrected_response else None,
            "feedback_type": f.feedback_type,
            "notes": f.notes,
            "category": f.category,
            "tags": json.loads(f.tags) if f.tags else [],
            "query_type": f.query_type,
            "similarity_score": f.similarity_score,
            "created_at": f.created_at.isoformat()
        } for f in feedbacks]
        
        session.close()
        return jsonify({'success': True, 'feedbacks': result}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/trainer/feedback/categories', methods=['GET'])
@token_required
def get_feedback_categories(user_id):
    """Obtener lista de categorías disponibles"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return jsonify({'error': 'Solo trainers'}), 403
        
        categories = session.query(TrainerFeedback.category)\
            .filter(TrainerFeedback.category.isnot(None))\
            .distinct()\
            .all()
        
        categories = [c[0] for c in categories]
        session.close()
        
        return jsonify({'success': True, 'categories': categories}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/trainer/upload', methods=['POST'])
@token_required
def upload_excel(user_id):
    """Subir Excel dinámico - crea tabla automáticamente"""
    session = None
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return jsonify({'error': 'Solo trainers pueden subir archivos'}), 403
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Solo archivos .xlsx'}), 400
        
        # Obtener nombre de tabla del archivo
        table_name = file.filename.replace('.xlsx', '').lower()
        table_name = table_name.replace(' ', '_')
        table_name = table_name.replace('-', '_')
        
        # Validar nombre válido
        if not table_name.isidentifier():
            return jsonify({'error': 'Nombre de archivo inválido'}), 400
        
        # Leer Excel
        excel_file = io.BytesIO(file.read())
        df = pd.read_excel(excel_file)
        
        logger.info(f"📊 Columnas del Excel: {list(df.columns)}")
        
        # NUEVO: Verificar si tabla ya existe ANTES de validar 'id'
        table_exists = False
        with engine.connect() as conn:
            stmt = text(f"SELECT 1 FROM information_schema.tables WHERE table_name = '{table_name}'")
            result = conn.execute(stmt).first()
            table_exists = result is not None
        
        # Validar que tenga columna 'id' SOLO si tabla es nueva
        if not table_exists:
            # Tabla nueva: EXIGE 'id'
            if 'id' not in df.columns:
                logger.error(f"❌ Falta columna 'id' en tabla nueva")
                return jsonify({'error': 'Excel debe tener columna "id" para tabla nueva'}), 400
        else:
            # Tabla existe: NO exige 'id'
            logger.info(f"✅ Tabla {table_name} existe, omitiendo validación de 'id'")
        
        # NUEVO: Forzar rollback de transacciones previas
        try:
            with engine.connect() as conn:
                conn.execute(text("ROLLBACK"))
        except:
            pass
        
        # Crear tabla o agregar datos
        try:
            with engine.connect() as conn:
                rows_inserted = 0
                errors = []
                
                if not table_exists:
                    # TABLA NUEVA: CREATE TABLE
                    logger.info(f"📊 Creando tabla nueva: {table_name}")
                    
                    columns_sql = "id SERIAL PRIMARY KEY"
                    
                    for col in df.columns:
                        if col != 'id':
                            if df[col].dtype in ['int64', 'int32']:
                                col_type = 'INTEGER'
                            elif df[col].dtype in ['float64', 'float32']:
                                col_type = 'NUMERIC'
                            else:
                                col_type = 'VARCHAR(255)'
                            
                            columns_sql += f", {col} {col_type}"
                    
                    create_sql = f"CREATE TABLE {table_name} ({columns_sql})"
                    conn.execute(text(create_sql))
                    conn.commit()
                    logger.info(f"✅ Tabla creada: {table_name}")
                else:
                    # TABLA EXISTE: Solo INSERT
                    logger.info(f"📊 Tabla {table_name} existe, agregando datos...")
                
                # Insertar datos (nuevo o existente)
                for idx, row in df.iterrows():
                    try:
                        cols = ", ".join([col for col in df.columns if col != 'id'])
                        placeholders = ", ".join([f"'{str(row[col]).replace(chr(39), chr(39)*2)}'" 
                                                 for col in df.columns if col != 'id'])
                        
                        insert_sql = f"INSERT INTO {table_name} ({cols}) VALUES ({placeholders})"
                        conn.execute(text(insert_sql))
                        rows_inserted += 1
                        log_audit(user_id, user.username, 'UPLOAD', details=f"Tabla: {table_name}, Filas: {rows_inserted}", ip_address=request.remote_addr)
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                
                conn.commit()
        
        except Exception as e:
            logger.error(f"Error en tabla: {e}")
            raise
        
        # Guardar metadata (solo si tabla nueva)
        if not table_exists:
            try:
                meta_session = Session()
                dynamic_table = DynamicTable(
                    table_name=table_name,
                    columns=json.dumps(list(df.columns)),
                    created_by=user_id
                )
                meta_session.add(dynamic_table)
                meta_session.commit()
                meta_session.close()
            except Exception as meta_err:
                logger.error(f"Error guardando metadata: {meta_err}")
        
        session.close()
        
        # Mensaje según si es tabla nueva o existente
        if table_exists:
            logger.info(f"✅ Tabla {table_name}: {rows_inserted} filas agregadas")
            message = f"Se agregaron {rows_inserted} filas a la tabla {table_name}"
        else:
            logger.info(f"✅ Tabla dinámica: {table_name} - {rows_inserted} filas")
            message = f"Tabla {table_name} creada con {rows_inserted} filas"
        
        # ENVIAR EMAIL EN AMBOS CASOS
        send_email_notification(table_name, list(df.columns), user.username, is_new_table=not table_exists)
        
        return jsonify({
            'success': True,
            'table_name': table_name,
            'rows_inserted': rows_inserted,
            'columns': list(df.columns),
            'errors': errors,
            'message': message,
            'table_exists': table_exists
        }), 200
        
    except Exception as e:
        logger.error(f"Error uploadando Excel: {e}")
        if session:
            session.close()
        return jsonify({'error': str(e)}), 500
    

@app.route('/api/export/chart', methods=['POST'])
@token_required
def export_chart(user_id):
    """Exportar gráfico como PNG"""
    try:
        data = request.json
        chart_type = data.get('chart_type')
        chart_data = data.get('chart_data')
        
        if not chart_type or not chart_data:
            return jsonify({'error': 'Datos incompletos'}), 400
        
        
        # Crear gráfico con Plotly
        if chart_type == 'bar':
            labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
            values = [row.get('facturacion', 0) for row in chart_data]
            
            fig = go.Figure(data=[go.Bar(x=labels, y=values, marker_color='#58a6ff')])
            fig.update_layout(
                title='Gráfico de Barras',
                xaxis_title='Clientes',
                yaxis_title='Facturación',
                template='plotly_dark',
                height=500
            )
        
        elif chart_type == 'pie':
            labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
            values = [row.get('market_share', 0) for row in chart_data]
            
            fig = go.Figure(data=[go.Pie(labels=labels, values=values)])
            fig.update_layout(
                title='Gráfico Circular',
                template='plotly_dark',
                height=500
            )
        
        elif chart_type == 'line':
            labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
            values = [row.get('facturacion', 0) for row in chart_data]
            
            fig = go.Figure(data=[go.Scatter(x=labels, y=values, mode='lines+markers', marker_color='#58a6ff')])
            fig.update_layout(
                title='Gráfico de Línea',
                xaxis_title='Clientes',
                yaxis_title='Facturación',
                template='plotly_dark',
                height=500
            )
        
        # Convertir a PNG
        img_bytes = fig.to_image(format="png")
        img_base64 = base64.b64encode(img_bytes).decode()
        
        return jsonify({
            'success': True,
            'image': img_base64,
            'filename': f'grafico_{chart_type}.png'
        }), 200
        
    except Exception as e:
        logger.error(f"Error exportando gráfico: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/pdf', methods=['POST'])
@token_required
def export_pdf(user_id):
    """Exportar respuesta + gráfico como PDF"""
    try:
        data = request.json
        response = data.get('response', '')
        chart_type = data.get('chart_type')
        chart_data = data.get('chart_data')
        
        # Crear PDF en memoria
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor='#58a6ff',
            spaceAfter=12
        )
        
        # Título
        elements.append(Paragraph('JARVIS - Reporte', title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Fecha
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        elements.append(Paragraph(f'<b>Fecha:</b> {fecha}', styles['Normal']))
        elements.append(Spacer(1, 0.4*inch))
        
        # Respuesta
        elements.append(Paragraph(response, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Generar gráfico si existe
        if chart_type and chart_data:
            try:
                # Crear gráfico con Plotly
                if chart_type == 'bar':
                    labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
                    values = [row.get('facturacion', 0) for row in chart_data]
                    
                    fig = go.Figure(data=[go.Bar(x=labels, y=values, marker_color='#58a6ff')])
                    fig.update_layout(
                        title='Gráfico de Barras',
                        xaxis_title='Clientes',
                        yaxis_title='Facturación',
                        template='plotly_dark',
                        height=400,
                        width=600
                    )
                
                elif chart_type == 'pie':
                    labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
                    values = [row.get('market_share', 0) for row in chart_data]
                    
                    fig = go.Figure(data=[go.Pie(labels=labels, values=values)])
                    fig.update_layout(
                        title='Gráfico Circular',
                        template='plotly_dark',
                        height=400,
                        width=600
                    )
                
                elif chart_type == 'line':
                    labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
                    values = [row.get('facturacion', 0) for row in chart_data]
                    
                    fig = go.Figure(data=[go.Scatter(x=labels, y=values, mode='lines+markers', marker_color='#58a6ff')])
                    fig.update_layout(
                        title='Gráfico de Línea',
                        xaxis_title='Clientes',
                        yaxis_title='Facturación',
                        template='plotly_dark',
                        height=400,
                        width=600
                    )
                
                # Convertir a PNG
                img_bytes = fig.to_image(format="png")
                img_buffer = BytesIO(img_bytes)
                
                # Agregar imagen al PDF
                elements.append(Spacer(1, 0.2*inch))
                elements.append(RLImage(img_buffer, width=5.5*inch, height=3.3*inch))
                
            except Exception as e:
                logger.error(f"Error generando gráfico en PDF: {e}")
                elements.append(Paragraph(f'<i>Error al incluir gráfico: {str(e)}</i>', styles['Normal']))
        
        # Generar PDF
        doc.build(elements)
        pdf_buffer.seek(0)
        pdf_base64 = base64.b64encode(pdf_buffer.read()).decode()
        
        return jsonify({
            'success': True,
            'pdf': pdf_base64,
            'filename': f'reporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        }), 200
        
    except Exception as e:
        logger.error(f"Error exportando PDF: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel', methods=['POST'])
@token_required
def export_excel(user_id):
    """Exportar datos como Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        import base64
        from datetime import datetime
        
        data = request.json
        table_data = data.get('data', [])
        filename = data.get('filename', 'datos')
        
        if not table_data:
            return jsonify({'error': 'Sin datos para exportar'}), 400
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Headers
        headers = list(table_data[0].keys()) if table_data else []
        header_fill = PatternFill(start_color='58a6ff', end_color='58a6ff', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, row_data in enumerate(table_data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = row_data.get(header, '')
                cell.alignment = Alignment(horizontal='left')
        
        # Auto-ajustar ancho
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_base64 = base64.b64encode(excel_buffer.read()).decode()
        
        return jsonify({
            'success': True,
            'excel': excel_base64,
            'filename': f'{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        }), 200
        
    except Exception as e:
        logger.error(f"Error exportando Excel: {e}")
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/audit/logs', methods=['GET'])
@token_required
def get_audit_logs(user_id):
    """Obtener logs de auditoría"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return jsonify({'error': 'Solo trainers pueden ver auditoría'}), 403
        
        logs = session.query(AuditLog)\
            .order_by(AuditLog.created_at.desc())\
            .limit(500)\
            .all()
        
        result = [{
            "id": log.id,
            "user_id": log.user_id,
            "username": log.username,
            "action": log.action,
            "details": log.details,
            "ip_address": log.ip_address,
            "created_at": log.created_at.isoformat()
        } for log in logs]
        
        session.close()
        return jsonify({'success': True, 'logs': result}), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo logs: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/trainer/tables', methods=['GET'])
@token_required

def get_tables(user_id):
    """Obtener lista de tablas dinámicas creadas por usuarios"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return jsonify({'error': 'Solo trainers'}), 403
        
        # Obtener solo tablas dinámicas creadas por usuarios (NO del sistema)
        tables_from_db = session.query(DynamicTable)\
            .filter(~DynamicTable.table_name.in_([
                'users', 'conversations', 'trainer_feedback', 'dynamic_tables', 
                'audit_logs', 'dim_anunciante', 'dim_anunciante_alias', 
                'dim_anunciante_perfil', 'fac_facturacion', 'jarvis_query_logs'
            ]))\
            .all()
        
        # Verificar que las tablas existan en la BD (sincronizar)
        valid_tables = []
        for t in tables_from_db:
            try:
                with engine.connect() as conn:
                    stmt = text(f"SELECT 1 FROM information_schema.tables WHERE table_name = '{t.table_name}'")
                    exists = conn.execute(stmt).first()
                    if exists:
                        valid_tables.append(t)
                    else:
                        # Tabla no existe, eliminarla de DynamicTable
                        logger.info(f"🗑️ Tabla {t.table_name} no existe en BD, eliminando de metadata")
                        session.query(DynamicTable).filter_by(table_name=t.table_name).delete()
                        session.commit()
            except Exception as sync_err:
                logger.error(f"Error sincronizando tabla {t.table_name}: {sync_err}")
        
        result = [{
            "table_name": t.table_name,
            "columns": json.loads(t.columns) if t.columns else [],
            "created_at": t.created_at.isoformat() if t.created_at else ""
        } for t in valid_tables]
        
        session.close()
        return jsonify({'success': True, 'tables': result}), 200
    except Exception as e:
        logger.error(f"Error obteniendo tablas: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/trainer/export-template', methods=['POST'])
@token_required
def export_template(user_id):
    """Exportar template de tabla"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        import base64
        
        user_session = Session()
        user = user_session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            user_session.close()
            return jsonify({'error': 'Solo trainers'}), 403
        
        data = request.json
        table_name = data.get('table_name')
        
        if not table_name:
            return jsonify({'error': 'table_name requerido'}), 400
        
        # Obtener columnas de la tabla de PostgreSQL
        with engine.connect() as conn:
            stmt = text(f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = '{table_name}'
                ORDER BY ordinal_position
            """)
            result = conn.execute(stmt)
            columns = [row[0] for row in result]
        
        user_session.close()
        
        # Omitir 'id' del template
        columns = [col for col in columns if col != 'id']
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = table_name
        
        # Headers
        header_fill = PatternFill(start_color='58a6ff', end_color='58a6ff', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-ajustar ancho
        for col in ws.columns:
            max_length = len(str(col[0].value)) if col[0].value else 0
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 15)
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_base64 = base64.b64encode(excel_buffer.read()).decode()
        
        return jsonify({
            'success': True,
            'excel': excel_base64,
            'filename': f'{table_name}_template.xlsx'
        }), 200
        
    except Exception as e:
        logger.error(f"Error exportando template: {e}")
        return jsonify({'error': str(e)}), 500

def send_email_notification(table_name, columns, username, is_new_table=True):
    """Enviar email de notificación de nueva tabla o datos agregados"""
    try:
        sender_email = "danilo.sosa@texo.com.py"
        sender_password = "yfvp aiws uorn ycsr"
        receiver_email = "danilo.sosa@texo.com.py"
        
        # Cambiar título y contenido según si es tabla nueva o INSERT
        if is_new_table:
            subject = f"🔔 Nueva tabla creada: {table_name}"
            title = "Nueva Tabla Creada en JARVIS"
            action = "⚠️ Acción requerida: Integra esta tabla en JARVIS si es necesario."
        else:
            subject = f"📊 Datos agregados a tabla: {table_name}"
            title = "Datos Agregados a Tabla en JARVIS"
            action = "ℹ️ Información: Se han agregado datos a la tabla existente."
        
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = sender_email
        message["To"] = receiver_email
        
        html = f"""\
        <html>
          <body style="font-family: Arial, sans-serif;">
            <h2 style="color: #58a6ff;">{title}</h2>
            <table style="border-collapse: collapse; width: 100%;">
              <tr style="background: #21262d;">
                <td style="padding: 8px; border: 1px solid #30363d;"><strong>Tabla:</strong></td>
                <td style="padding: 8px; border: 1px solid #30363d;">{table_name}</td>
              </tr>
              <tr>
                <td style="padding: 8px; border: 1px solid #30363d;"><strong>Columnas:</strong></td>
                <td style="padding: 8px; border: 1px solid #30363d;">{', '.join(columns)}</td>
              </tr>
              <tr style="background: #21262d;">
                <td style="padding: 8px; border: 1px solid #30363d;"><strong>Usuario:</strong></td>
                <td style="padding: 8px; border: 1px solid #30363d;">{username}</td>
              </tr>
              <tr>
                <td style="padding: 8px; border: 1px solid #30363d;"><strong>Fecha:</strong></td>
                <td style="padding: 8px; border: 1px solid #30363d;">{datetime.now().strftime('%d/%m/%Y %H:%M')}</td>
              </tr>
            </table>
            <hr style="border: none; border-top: 1px solid #30363d; margin: 20px 0;">
            <p><strong>{action}</strong></p>
            <p>Cuando esté lista, notifica al usuario con este mensaje:</p>
            <p style="background: #21262d; padding: 10px; border-radius: 4px;"><code>✅ Tabla '{table_name}' ya está disponible para consultas en JARVIS</code></p>
          </body>
        </html>
        """
        
        part = MIMEText(html, "html")
        message.attach(part)
        
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, receiver_email, message.as_string())
        
        logger.info(f"✅ Email enviado: tabla {table_name}")
        return True
    except Exception as e:
        logger.error(f"❌ Error enviando email: {e}")
        return False
        logger.info(f"📧 Intentando enviar email...")
        logger.info(f"De: {sender_email}")
        logger.info(f"Para: {receiver_email}")
        logger.info(f"Tabla: {table_name}")

if __name__ == '__main__':
    logger.info("🚀 JARVIS Backend + Tablas Dinámicas iniciando...")
    app.run(host='0.0.0.0', port=5000, debug=True)