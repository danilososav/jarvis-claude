"""
JARVIS - Backend Flask con PostgreSQL + Autenticación + Fuzzy Matching + Tablas Dinámicas
Sistema BI conversacional para agencia de medios (Paraguay)
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from dotenv import load_dotenv
import logging
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
from functools import wraps
from fuzzywuzzy import fuzz
import pandas as pd
import io
import json
from decimal import Decimal
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage
from reportlab.lib.units import inch
from io import BytesIO
import base64
from datetime import datetime
import plotly.graph_objects as go
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from validate_feedback import validate_feedback_with_claude, format_validation_for_trainer
from busqueda_flexible import (
    get_facturacion_cliente,
    get_inversion_medios_cliente,
    get_ranking_dnit_cliente,
    get_perfil_adlens_cliente
)

try:
    from jarvis_360_integration import (
        get_cliente_360,
        identify_cliente_fuzzy_360,
        format_data_for_claude_360
    )
    # ✅ NO IMPORTAR engine - ya está definido en este mismo archivo

    print("✅ jarvis_360_integration importado correctamente")

except ImportError as e:
    print(f"❌ Error importando jarvis_360_integration: {e}")


# SQLAlchemy
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime
from sqlalchemy.orm import declarative_base, sessionmaker
from sqlalchemy import text

import base64
from io import BytesIO
import plotly.graph_objects as go
from claude_handler_v2 import ClaudeHandler
from chart_utils import (
    detect_query_intent, 
    build_chart_config,
    should_include_chart,
    should_include_text,
    build_table_config,
    build_kpi_config
)


load_dotenv()

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Inicializar Claude Handler
ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY')
claude_handler = ClaudeHandler(ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else None
if claude_handler:
    logger.info("✅ Claude Sonnet 4.5 inicializado")
else:
    logger.warning("⚠️ ANTHROPIC_API_KEY no encontrada - usando respuestas mock")

# Flask app
app = Flask(__name__)
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})


# ==================== DATABASE ====================

DB_USER = os.getenv('PG_USER', 'postgres')
DB_PASS = os.getenv('PG_PASS', '12345')
DB_HOST = os.getenv('PG_HOST', 'localhost')
DB_PORT = os.getenv('PG_PORT', '5432')
DB_NAME = os.getenv('PG_DB', 'jarvis')

DB_URL = f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(DB_URL, pool_pre_ping=True)

# SQLAlchemy models
Base = declarative_base()

class User(Base):
    __tablename__ = 'users'
    id = Column(Integer, primary_key=True)
    username = Column(String(50), unique=True)
    password_hash = Column(String(255))
    role = Column(String(20), default='normal')
    created_at = Column(DateTime, default=datetime.now)

class Conversation(Base):
    __tablename__ = 'conversations'
    id = Column(Integer, primary_key=True)
    user_id = Column(Integer)
    session_id = Column(String(50))  # NUEVO: agrupar chats
    query = Column(Text)
    response = Column(Text)
    query_type = Column(String(50))
    chart_config = Column(Text)
    chart_data = Column(Text)
    created_at = Column(DateTime, default=datetime.now)

class TrainerFeedback(Base):
    __tablename__ = 'trainer_feedback'
    id = Column(Integer, primary_key=True)
    conversation_id = Column(Integer)
    user_id = Column(Integer)
    original_query = Column(Text)
    original_response = Column(Text)
    corrected_response = Column(Text)
    feedback_type = Column(String(50))
    notes = Column(Text)
    category = Column(String(50))  # NUEVO
    tags = Column(Text)  # NUEVO - JSON array
    chart_config = Column(Text)  # NUEVO - JSON config
    query_type = Column(String(50))  # NUEVO
    similarity_score = Column(Integer)  # NUEVO
    # Columnas de validación
    status = Column(String(20), default='pending')  # NUEVO
    validation_verdict = Column(String(20))  # NUEVO
    validation_reasoning = Column(Text)  # NUEVO
    validation_date = Column(DateTime)  # NUEVO
    escalation_reason = Column(Text)  # NUEVO
    escalation_date = Column(DateTime)  # NUEVO
    resolved_by = Column(String(50))  # NUEVO
    resolution = Column(Text)  # NUEVO
    resolution_date = Column(DateTime)  # NUEVO
    created_at = Column(DateTime, default=datetime.now)

class DynamicTable(Base):
    __tablename__ = 'dynamic_tables'
    id = Column(Integer, primary_key=True)
    table_name = Column(String(100), unique=True)
    columns = Column(Text)
    created_by = Column(Integer)
    created_at = Column(DateTime, default=datetime.now)

class AuditLog(Base):
    __tablename__ = 'audit_logs'
    id = Column(Integer, primary_key=True)
    user_id = Column(Integer)
    username = Column(String(50))
    action = Column(String(100))
    details = Column(Text)
    ip_address = Column(String(45))
    created_at = Column(DateTime, default=datetime.now)

# Create tables
try:
    Base.metadata.create_all(engine)
    logger.info("✅ Tablas creadas/verificadas")
except Exception as e:
    logger.error(f"Error creando tablas: {e}")

Session = sessionmaker(bind=engine)

# ==================== AUTHENTICATION ====================

SECRET_KEY = os.getenv('SECRET_KEY', 'jarvis-secret-key-2026')

def generate_token(user_id):
    """Genera JWT token"""
    payload = {
        'user_id': user_id,
        'exp': datetime.utcnow() + timedelta(days=7)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

def verify_token(token):
    """Verifica JWT token"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        return payload['user_id']
    except:
        return None

def token_required(f):
    """Decorator para proteger endpoints"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return safe_jsonify({'error': 'Token requerido'}), 401
        
        token = token.replace('Bearer ', '')
        user_id = verify_token(token)
        
        if not user_id:
            return safe_jsonify({'error': 'Token inválido'}), 401
        
        return f(user_id, *args, **kwargs)
    return decorated

# -------------------------------------------------------
def safe_jsonify(data):
    """
    Versión segura de jsonify que maneja Decimals
    """
    def convert_decimals(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, dict):
            return {k: convert_decimals(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [convert_decimals(v) for v in obj]
        else:
            return obj
    
    safe_data = convert_decimals(data)
    return jsonify(safe_data)

# ==================== FUZZY MATCHING ====================

def log_audit(user_id, username, action, details='', ip_address=''):
    """Registra auditoría"""
    try:
        session = Session()
        audit = AuditLog(
            user_id=user_id,
            username=username,
            action=action,
            details=details,
            ip_address=ip_address
        )
        session.add(audit)
        session.commit()
        session.close()
    except Exception as e:
        logger.error(f"Error en auditoría: {e}")

def find_similar_feedback(user_query, category=None):
    """Busca feedback similar con threshold dinámico y categoría - SOLO APROBADOS"""
    try:
        session = Session()
        
        # SOLO buscar feedback aprobado
        approved_statuses = ['auto_approved', 'manually_approved']
        
        # Si hay categoría, filtrar por esa primero
        if category:
            feedbacks = session.query(TrainerFeedback)\
                .filter(TrainerFeedback.category == category)\
                .filter(TrainerFeedback.status.in_(approved_statuses))\
                .all()
        else:
            feedbacks = session.query(TrainerFeedback)\
                .filter(TrainerFeedback.status.in_(approved_statuses))\
                .all()
        
        session.close()
        
        best_match = None
        best_score = 0
        
        # Threshold dinámico según longitud de query
        query_len = len(user_query.split())
        if query_len <= 3:
            threshold = 75  # Queries cortas: menos exigentes
        elif query_len <= 7:
            threshold = 80  # Queries medianas
        else:
            threshold = 85  # Queries largas: más exigentes
        
        for fb in feedbacks:
            if fb.original_query:
                # Usar token_set_ratio para más flexibilidad
                score = fuzz.token_set_ratio(user_query.lower(), fb.original_query.lower())
                if score > best_score:
                    best_score = score
                    best_match = fb
        
        if best_match and best_score >= threshold:
            logger.info(f"✅ Feedback similar encontrado: score {best_score}, categoría: {best_match.category}")
            best_match.similarity_score = best_score
            return best_match
        
        logger.info(f"❌ No hay feedback similar (mejor score: {best_score}, threshold: {threshold})")
        return None
    except Exception as e:
        logger.error(f"Error en fuzzy matching: {e}")
        return None

# ==================== QUERIES A LA BD ====================
def format_data_for_claude(rows, query_type):
    """
    Formatea datos cruzados para que Claude los entienda mejor
    """
    if not rows:
        return rows
    
    # Si tiene datos de inversión o ranking, formatear especialmente
    if query_type == "facturacion" and isinstance(rows[0], dict):
        # Datos cruzados, formatear para Claude
        formatted = []
        for row in rows:
            item = {
                'cliente': row.get('cliente', ''),
                'facturacion': row.get('facturacion_total', 0), 
                'promedio_mensual': row.get('promedio_mensual', 0),
                'market_share': row.get('market_share', 0)
            }
            
            # Agregar inversión si existe
            if 'inversion_medios' in row:
                inv = row['inversion_medios']
                item['inversion_detalle'] = inv
                item['inversion_total_usd'] = sum(i.get('inversion_usd', 0) for i in inv)
            else:
                # ✅ Valores por defecto cuando no hay inversión
                item['inversion_detalle'] = []
                item['inversion_total_usd'] = 0
            
            # Agregar ranking si existe
            if 'ranking_dnit' in row:
                item['ranking'] = row['ranking_dnit'].get('ranking')
                item['aporte_dnit'] = row['ranking_dnit'].get('aporte_gs')
            else:
                # ✅ Valores por defecto cuando no hay ranking
                item['ranking'] = None
                item['aporte_dnit'] = 0
            
            formatted.append(item)
        
        return formatted
    
    return rows


def generate_response(query_type, rows, user_query):
    """
    Genera respuesta usando Claude API si está disponible, 
    sino usa respuesta mock básica
    """
    
    # Si hay respuesta corregida por trainer, usarla directamente
    if query_type == "corrected":
        return "Respuesta corregida por entrenador"
    
    # Intentar usar Claude API primero
    if claude_handler and rows:
        try:
            response = claude_handler.enhance_response(
                user_query=user_query,
                data=rows,
                query_type=query_type
            )
            if response:
                return response
        except Exception as e:
            logger.error(f"❌ Error usando Claude: {e}")
    
    # Fallback a respuestas mock si Claude no está disponible
    return _mock_response_fallback(query_type, rows, user_query)


def _mock_response_fallback(query_type, rows, user_query):
    """Respuestas básicas cuando Claude API no está disponible"""
    
    if query_type == "dynamic_table":
        if rows:
            return f"Encontré {len(rows)} registros en esa tabla. Primeros registros: {str(rows[:2])}"
        return "No hay datos en esa tabla."
    
    elif query_type == "chart":
        if rows:
            total = sum(r.get("facturacion", 0) for r in rows)
            return f"Aquí te muestro un gráfico con los {len(rows)} clientes principales. Facturación total: {total:,.0f} Gs"
        return "No hay datos disponibles para el gráfico."
    
    elif query_type == "ranking":
        if rows:
            total_facturacion = sum(r.get("facturacion", 0) for r in rows)
            parts = [f"{i+1}. {r['cliente']}: {r['facturacion']:,.0f} Gs ({r.get('market_share', 0):.2f}%)" 
                    for i, r in enumerate(rows)]
            return f"Top {len(rows)} clientes por facturación:\n" + "\n".join(parts) + f"\nTotal: {total_facturacion:,.0f} Gs"
        return "No encontré datos de clientes."
    
    elif query_type == "facturacion":
        if rows:
            r = rows[0]
            return f"**{r['cliente']}** facturó **{r['facturacion']:,.0f} Gs** con un market share de {r.get('market_share', 0):.2f}%. Promedio mensual: {r.get('promedio_mensual', 0):,.0f} Gs."
        return "No tengo datos de facturación para ese cliente."
    
    else:
        return "Consulta procesada."

def get_top_clientes_enriched(query):
    """Top clientes por facturación - VERSIÓN FINAL CORRECTA"""
    try:
        with engine.connect() as conn:
            stmt = text("""
                SELECT 
                    cliente_original,
                    COUNT(*) as registros,
                    SUM(facturacion) as facturacion_total
                FROM fact_facturacion
                WHERE cliente_original IS NOT NULL
                GROUP BY cliente_original
                ORDER BY facturacion_total DESC
                LIMIT 10
            """)
            rows = conn.execute(stmt).fetchall()
            
            # Calcular total para market share
            total_stmt = text("SELECT SUM(facturacion) FROM fact_facturacion WHERE cliente_original IS NOT NULL")
            total_result = conn.execute(total_stmt).fetchone()
            total_facturacion = float(total_result[0]) if total_result[0] else 1
            
            result = []
            for r in rows:
                facturacion = float(r[2]) if r[2] else 0
                market_share = (facturacion / total_facturacion * 100) if total_facturacion > 0 else 0
                
                result.append({
                    "cliente": r[0] or "Sin nombre",
                    "facturacion": facturacion,
                    "registros": int(r[1]) if r[1] else 0,
                    "market_share": market_share
                })
            return result
    except Exception as e:
        logger.error(f"Error Top Clientes: {e}")
        return []
    

    # metodo que busca si un cliente tiene busqueda de facturacion o inversion separados o juntos.
    
def get_facturacion_enriched(query):
    """
    Facturación de un cliente con búsqueda flexible
    Detecta si pide solo facturación o facturación + inversión
    """
    query_limpio = query.replace('?', '').replace('!', '').replace(',', '').lower()
    
    # Palabras comunes a ignorar
    stopwords = ['cuanto', 'cuánto', 'facturo', 'facturó', 'facturacion', 'facturación', 
                 'de', 'la', 'el', 'en', 'y', 'o', 'para', 'con', 'a', 'un', 'una',
                 'invirtio', 'invirti', 'invertir', 'inversion']
    
    palabras = query_limpio.split()
    cliente_palabras = [p for p in palabras if p not in stopwords and len(p) > 2]
    
    if not cliente_palabras:
        return []
    
    # Extraer nombre del cliente
    cliente = cliente_palabras[0] if len(cliente_palabras) == 1 else " ".join(cliente_palabras[:2])
    
    # Detectar qué datos necesita
    pide_inversion = any(w in query_limpio for w in ['tv', 'radio', 'cable', 'inversion', 'invirtio', 'invertir', 'medios', 'publicidad', 'pauta'])
    pide_ranking = any(w in query_limpio for w in ['ranking', 'dnit', 'posicion', 'puesto', 'aporte'])
    
    resultado = []
    
    try:
        # 1. Siempre buscar facturación si la query lo indica
        if any(w in query_limpio for w in ['facturo', 'facturacion', 'vendio', 'ventas', 'cuanto']):
            facturacion = get_facturacion_cliente(cliente, engine)
            
            if facturacion:
                # Agregar datos de facturación al resultado
                for f in facturacion:
                    resultado.append(f)
        
        # 2. Si pide inversión, buscar en medios
        if pide_inversion:
            # Detectar filtros
            filtros = {}
            if 'tv' in query_limpio:
                filtros['medio'] = 'TV'
            elif 'radio' in query_limpio:
                filtros['medio'] = 'RADIO'
            elif 'cable' in query_limpio:
                filtros['medio'] = 'CABLE'
            
            inversion = get_inversion_medios_cliente(cliente, engine, filtros)
            
            if inversion:
                # Agregar datos de inversión al resultado
                if resultado:
                    # Ya hay facturación, agregar inversión al mismo dict
                    resultado[0]['inversion_medios'] = inversion
                else:
                    # Solo inversión, crear resultado
                    resultado = [{
                        'cliente': inversion[0]['cliente'],
                        'inversion_medios': inversion
                    }]
        
        # 3. ✅ CORREGIDO: Siempre buscar ranking DNIT cuando tenemos datos del cliente
        if resultado:  # ✅ CORREGIDO: buscar automáticamente
            ranking = get_ranking_dnit_cliente(cliente, engine)
            
            if ranking:
                resultado[0]['ranking_dnit'] = ranking[0]
        
        return resultado
        
    except Exception as e:
        logger.error(f"Error en get_facturacion_enriched: {e}")
        return []

# ==================== ENDPOINTS ====================

@app.route('/api/health', methods=['GET'])
def health():
    """Health check"""
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return safe_jsonify({"status": "✅ OK", "db": "connected"}), 200
    except:
        return safe_jsonify({"status": "❌ ERROR", "db": "disconnected"}), 500

# ==================== AUTHENTICATION ENDPOINTS ====================

@app.route('/api/auth/register', methods=['POST'])
def register():
    """Registrar nuevo usuario"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return safe_jsonify({'error': 'Username y password requeridos'}), 400
        
        if len(password) < 6:
            return safe_jsonify({'error': 'Password mínimo 6 caracteres'}), 400
        
        session = Session()
        
        existing = session.query(User).filter_by(username=username).first()
        if existing:
            session.close()
            return safe_jsonify({'error': 'Usuario ya existe'}), 409
        
        user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role='normal'
        )
        session.add(user)
        session.commit()
        user_id = user.id
        session.close()
        
        token = generate_token(user_id)
        
        logger.info(f"✅ Usuario registrado: {username}")
        
        return safe_jsonify({
            'success': True,
            'user_id': user_id,
            'username': username,
            'role': 'normal',
            'token': token
        }), 201
        
    except Exception as e:
        logger.error(f"Error registrando usuario: {e}")
        return safe_jsonify({'error': str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login de usuario"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return safe_jsonify({'error': 'Username y password requeridos'}), 400
        
        session = Session()
        user = session.query(User).filter_by(username=username).first()
        session.close()
        
        if not user or not check_password_hash(user.password_hash, password):
            return safe_jsonify({'error': 'Credenciales inválidas'}), 401
        
        token = generate_token(user.id)
        
        logger.info(f"✅ Login: {username}")
        
        return safe_jsonify({
            'success': True,
            'user_id': user.id,
            'username': user.username,
            'role': user.role,
            'token': token
        }), 200
        
        log_audit(user.id, user.username, 'LOGIN', ip_address=request.remote_addr)
    except Exception as e:
        logger.error(f"Error en login: {e}")
        return safe_jsonify({'error': str(e)}), 500

@app.route('/api/auth/verify', methods=['GET'])
@token_required
def verify_user(user_id):
    """Verificar token y obtener info del usuario"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        session.close()
        
        if not user:
            return safe_jsonify({'error': 'Usuario no encontrado'}), 404
        
        return safe_jsonify({
            'success': True,
            'user_id': user.id,
            'username': user.username,
            'role': user.role
        }), 200
        
    except Exception as e:
        return safe_jsonify({'error': str(e)}), 500

# ==================== QUERY ENDPOINTS ====================
@app.route('/api/chat/history', methods=['GET'])
@token_required
def get_chat_history(user_id):
    """Obtener historial agrupado por sesión"""
    try:
        session = Session()
        
        # Obtener todas las conversaciones
        conversations = session.query(Conversation)\
            .filter_by(user_id=user_id)\
            .order_by(Conversation.created_at.desc())\
            .all()
        
        # Agrupar por session_id
        grouped = {}
        for c in conversations:
            sid = c.session_id or f"session_{user_id}"
            if sid not in grouped:
                grouped[sid] = []
            grouped[sid].append({
                "id": c.id,
                "query": c.query,
                "response": c.response,
                "query_type": c.query_type,
                "chart_config": json.loads(c.chart_config) if c.chart_config else None,
                "rows": json.loads(c.chart_data) if c.chart_data else None,
                "created_at": c.created_at.isoformat()
            })
        
        # Convertir a lista de sesiones
        result = []
        for session_id, messages in grouped.items():
            result.append({
                "session_id": session_id,
                "messages": messages,
                "created_at": messages[0]["created_at"]
            })
        
        session.close()
        return safe_jsonify({"success": True, "sessions": result}), 200
    except Exception as e:
        logger.error(f"Error obteniendo historial: {e}")
        return safe_jsonify({"error": str(e)}), 500
    
@app.route('/api/chat/history/<int:conv_id>', methods=['DELETE'])
@token_required
def delete_conversation(user_id, conv_id):
    """Eliminar una conversación"""
    session = None
    try:
        session = Session()
        
        conversation = session.query(Conversation).filter_by(
            id=conv_id, 
            user_id=user_id
        ).first()
        
        if not conversation:
            session.close()
            return safe_jsonify({'error': 'Conversación no encontrada'}), 404
        
        session.delete(conversation)
        session.commit()
        session.close()
        
        logger.info(f"✅ Conversación {conv_id} eliminada")
        
        return safe_jsonify({'success': True}), 200
        
    except Exception as e:
        logger.error(f"Error eliminando conversación: {e}")
        if session:
            session.rollback()
            session.close()
        return safe_jsonify({'error': str(e)}), 500

@app.route('/api/query', methods=['POST'])
@token_required
def query(user_id):
    """Procesar query con sistema de gráficos y texto separados"""
    data = request.json
    user_query = data.get('query', '').strip()
    session_id = data.get('session_id', str(user_id))
    
    session = Session()
    user = session.query(User).filter_by(id=user_id).first()
    username = user.username if user else 'unknown'
    session.close()
    
    log_audit(user_id, username, 'QUERY', details=user_query[:100], ip_address=request.remote_addr)

    if not user_query:
        return safe_jsonify({"error": "Query vacío"}), 400
    
    try:
        query_lower = user_query.lower()
        
        # Detectar intención: chart_only, text_only, o chart_and_text
        intent = detect_query_intent(user_query)
        logger.info(f"🎯 Intención detectada: {intent}")
        
        # BUSCAR FEEDBACK SIMILAR PRIMERO (prioridad máxima)
        similar_feedback = find_similar_feedback(user_query)
        if similar_feedback:
            response_text = similar_feedback.corrected_response
            return safe_jsonify({
                "success": True,
                "responses": [{
                    "type": "text",
                    "content": response_text,
                    "query_type": "corrected"
                }]
            }), 200
        
        # DETECCIÓN DE TABLAS DINÁMICAS
        dynamic_session = Session()
        try:
            dynamic_tables = dynamic_session.query(DynamicTable).all()
            for dt in dynamic_tables:
                if dt.table_name in query_lower:
                    with engine.connect() as conn:
                        stmt = text(f"SELECT * FROM {dt.table_name} LIMIT 10")
                        result = conn.execute(stmt).fetchall()
                        rows = [dict(row._mapping) for row in result]
                        
                    response_text = f"Encontré {len(rows)} registros en la tabla {dt.table_name}"
                    return safe_jsonify({
                        "success": True,
                        "responses": [{
                            "type": "text",
                            "content": response_text,
                            "query_type": "dynamic_table",
                            "data": rows
                        }]
                    }), 200
        finally:
            dynamic_session.close()
        
        # DETECCIÓN DE TIPO DE QUERY
        rows = []
        query_type = "generico"

        # 1. RANKINGS
        if any(w in query_lower for w in ["top", "ranking", "principal", "importante", "mayor", "más", "clientes"]):
            query_type = "ranking"
            rows = get_top_clientes_enriched(user_query)
            logger.info(f"🔍 Detectado: ranking - rows: {len(rows)}")

        # 2. FACTURACIÓN CON KEYWORDS
        elif any(w in query_lower for w in ["cuánto", "cuanto", "factur", "how much", "invirti", "ranking", "dnit", "datos", "perfil", "informacion", "cluster", "cultura"]):
            query_type = "facturacion"
            rows = get_cliente_360(user_query, engine)
            rows = format_data_for_claude_360(rows, query_type)
            logger.info(f"🔍 Detectado: facturacion 360° con keywords - rows: {len(rows)}")

       # 3. ✅ CONSULTAS COMPLEJAS SIN CLAUDE
        elif any(w in query_lower for w in ["comparar", "compare", "vs", "ranking", "clusters", "analisis", "estadisticas", "mercado", "completo", "datos completos"]):
            from jarvis_360_integration import query_compleja_sin_claude_handler_puro
            complex_result, status_code = query_compleja_sin_claude_handler_puro(user_query, engine)
            return safe_jsonify(complex_result), status_code
        
        else:
            from jarvis_360_integration import es_consulta_de_cliente
            cliente_detectado = es_consulta_de_cliente(user_query, engine)
            
            if cliente_detectado:
                query_type = "facturacion"
                rows = get_cliente_360(user_query, engine)
                rows = format_data_for_claude_360(rows, query_type)
                logger.info(f"🔍 Detectado: cliente automático '{cliente_detectado['nombre']}' - rows: {len(rows)}")
            else:
                return safe_jsonify({
                    "success": False, 
                    "response": "Consulta no reconocida. Prueba: 'Top 5 clientes', 'Cuánto facturó CERVEPAR?', o simplemente el nombre de un cliente como 'Unilever'"
                }), 200
        
        # ✅ CONSTRUIR RESPUESTAS SEGÚN INTENCIÓN (LIMPIO)
        responses = []
        
        if intent == "table_only":
            # Solo tabla
            table_config = build_table_config(query_type, rows, user_query)
            responses.append({
                "type": "table",
                "table_config": table_config,
                "data": rows,
                "query_type": query_type
            })
        
        elif intent == "kpi_only":
            # Solo KPI card
            kpi_config = build_kpi_config(query_type, rows, user_query)
            responses.append({
                "type": "kpi",
                "kpi_config": kpi_config,
                "data": rows,
                "query_type": query_type
            })
        
        elif intent == "chart_and_text":
            # Mensaje informativo
            responses.append({
                "type": "info",
                "content": "📊 Te muestro el gráfico y el análisis por separado"
            })
            
            # Gráfico
            chart_config = build_chart_config(user_query, query_type, rows)
            responses.append({
                "type": "chart",
                "chart_config": chart_config,
                "data": rows,
                "query_type": query_type
            })
            
            # ✅ Análisis de Claude (UNA SOLA VEZ)
            analysis_text = generate_response(query_type, rows, user_query)
            responses.append({
                "type": "text",
                "content": analysis_text,
                "query_type": query_type,
                "data": rows
            })
        
        elif intent == "chart_only":
            # Solo gráfico
            chart_config = build_chart_config(user_query, query_type, rows)
            responses.append({
                "type": "chart",
                "chart_config": chart_config,
                "data": rows,
                "query_type": query_type
            })
        
        else:  # text_only
            # ✅ Solo análisis de Claude
            analysis_text = generate_response(query_type, rows, user_query)
            responses.append({
                "type": "text",
                "content": analysis_text,
                "query_type": query_type,
                "data": rows
            })
        
        # ✅ GUARDAR EN BD (solo la primera respuesta para historial)
        # ✅ GUARDAR EN BD (solo la primera respuesta para historial)
        try:            
            # Validación mejorada
            if responses and len(responses) > 0:
                main_response = responses[0]
            else:
                main_response = {"type": "text", "content": "Sin respuesta"}
            
            conversation = Conversation(
                user_id=user_id,
                session_id=session_id,
                query=user_query,
                response=main_response.get("content", ""),
                query_type=query_type,
                chart_config=json.dumps(main_response.get("chart_config")) if main_response.get("chart_config") else None,
                chart_data=json.dumps(rows, default=float) if rows else None
            )
            
            # ✅ USAR MISMO PATRÓN QUE TRAINER FEEDBACK
            bd_session = Session()
            bd_session.add(conversation)
            bd_session.commit()
            conversation_id = conversation.id

            logger.info(f"✅ Conversación guardada: {conversation_id}")
            
        except Exception as bd_error:
            logger.error(f"❌ Error guardando en BD: {bd_error}")
            if bd_session:
                bd_session.rollback()
        finally:
            if bd_session:
                bd_session.close()
            
        return safe_jsonify({
            "success": True,
            "responses": responses
        }), 200
        
    except Exception as e:
        logger.error(f"Error en query: {e}")
        return safe_jsonify({"error": "Error procesando consulta"}), 500

# ==================== TRAINER ENDPOINTS ====================
@app.route('/api/trainer/feedback', methods=['POST'])
@token_required
def submit_trainer_feedback(user_id):
    """Enviar corrección/feedback como trainer con validación automática de Claude"""
    session = None
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        username = user.username if user else 'unknown'
        
        if not user or user.role != 'trainer':
            session.close()
            return safe_jsonify({'error': 'Solo trainers pueden enviar feedback'}), 403
        
        data = request.json
        
        # Crear feedback con status pendiente
        feedback = TrainerFeedback(
            conversation_id=data.get('conversation_id'),
            user_id=user_id,
            original_query=data.get('original_query'),
            original_response=data.get('original_response'),
            corrected_response=data.get('corrected_response'),
            feedback_type=data.get('feedback_type', 'correction'),
            notes=data.get('notes', ''),
            category=data.get('category'),
            tags=json.dumps(data.get('tags', [])),
            chart_config=None,
            query_type=data.get('query_type'),
            status='pending'  # NUEVO
        )
        session.add(feedback)
        session.commit()
        feedback_id = feedback.id
        
        # Preparar datos para validación
        feedback_data = {
            'original_query': data.get('original_query'),
            'original_response': data.get('original_response'),
            'corrected_response': data.get('corrected_response'),
            'data_snapshot': data.get('data_snapshot', [])
        }
        
        # VALIDAR CON CLAUDE
        try:
            validation_result = validate_feedback_with_claude(feedback_data, claude_handler)
            
            # Actualizar feedback con resultado
            feedback.validation_verdict = validation_result['verdict']
            feedback.validation_reasoning = validation_result['reasoning']
            feedback.validation_date = datetime.now()
            
            if validation_result['verdict'] == 'approved':
                feedback.status = 'auto_approved'
            else:
                feedback.status = 'auto_rejected'
            
            session.commit()
            logger.info(f"✅ Feedback {feedback_id} validado: {validation_result['verdict']}")
            
        except Exception as validation_error:
            logger.error(f"❌ Error en validación: {validation_error}")
            
            feedback.status = 'auto_rejected'
            feedback.validation_verdict = 'rejected'
            feedback.validation_reasoning = f'Error en validación: {str(validation_error)}'
            feedback.validation_date = datetime.now()
            session.commit()
            
            validation_result = {
                'verdict': 'rejected',
                'trainer_message': 'Hubo un error técnico al validar tu corrección.',
                'reasoning': str(validation_error)
            }
        
        session.close()
        
        log_audit(user_id, username, 'FEEDBACK', 
                 details=f"Feedback ID: {feedback_id}, Verdict: {validation_result['verdict']}", 
                 ip_address=request.remote_addr)
        
        # Formatear respuesta para el trainer
        formatted_response = format_validation_for_trainer(validation_result)
        formatted_response['feedback_id'] = feedback_id  # AGREGAR feedback_id
        
        return safe_jsonify({
            'success': True,
            'feedback_id': feedback_id,
            'validation': formatted_response
        }), 201
        
    except Exception as e:
        logger.error(f"Error en trainer feedback: {e}")
        if session:
            session.rollback()
            session.close()
        return safe_jsonify({'error': str(e)}), 500

@app.route('/api/trainer/feedback/<int:feedback_id>/escalate', methods=['POST'])
@token_required
def escalate_feedback(user_id, feedback_id):
    """Escalar feedback cuando trainer no está de acuerdo con Claude"""
    session = None
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        username = user.username if user else 'unknown'
        
        if not user or user.role != 'trainer':
            session.close()
            return safe_jsonify({'error': 'Solo trainers pueden escalar feedback'}), 403
        
        # Buscar feedback
        feedback = session.query(TrainerFeedback).filter_by(id=feedback_id).first()
        
        if not feedback:
            session.close()
            return safe_jsonify({'error': 'Feedback no encontrado'}), 404
        
        # Verificar que esté rechazado
        if feedback.status != 'auto_rejected':
            session.close()
            return safe_jsonify({'error': 'Solo se pueden escalar correcciones rechazadas'}), 400
        
        data = request.json
        escalation_reason = data.get('reason', '')
        
        # Actualizar feedback
        feedback.status = 'escalated'
        feedback.escalation_reason = escalation_reason
        feedback.escalation_date = datetime.now()
        session.commit()
        
        # Enviar email al administrador
        try:
            send_escalation_email(feedback, user, escalation_reason)
            logger.info(f"📧 Email de escalación enviado para feedback {feedback_id}")
        except Exception as email_error:
            logger.error(f"❌ Error enviando email: {email_error}")
        
        session.close()
        
        log_audit(user_id, username, 'ESCALATE_FEEDBACK', 
                 details=f"Feedback ID: {feedback_id}", 
                 ip_address=request.remote_addr)
        
        return safe_jsonify({
            'success': True,
            'message': 'Reporte enviado al administrador. Serás notificado cuando se resuelva.'
        }), 200
        
    except Exception as e:
        logger.error(f"Error escalando feedback: {e}")
        if session:
            session.rollback()
            session.close()
        return safe_jsonify({'error': str(e)}), 500


def send_escalation_email(feedback, trainer_user, escalation_reason):
    """Envía email al administrador cuando un trainer escala un feedback"""
    
    # Configuración SMTP
    smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    admin_email = os.getenv('ADMIN_EMAIL', 'danilo.sosa@texo.com.py')
    
    if not smtp_user or not smtp_pass:
        logger.warning("SMTP no configurado, email no enviado")
        return
    
    # Formatear datos
    data_preview = "N/A"
    try:
        if feedback.chart_data:
            data_json = json.loads(feedback.chart_data)
            data_preview = json.dumps(data_json[:3], indent=2, ensure_ascii=False)
    except:
        pass
    
    # Construir email
    subject = f"🚨 JARVIS - Trainer reporta desacuerdo (Feedback #{feedback.id})"
    
    body = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DESACUERDO REPORTADO
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Trainer: {trainer_user.username}
Fecha: {datetime.now().strftime('%d/%b/%Y %H:%M')}

QUERY ORIGINAL:
"{feedback.original_query}"

RESPUESTA ORIGINAL DE CLAUDE:
{feedback.original_response}

CORRECCIÓN PROPUESTA POR TRAINER:
{feedback.corrected_response}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
VALIDACIÓN DE CLAUDE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Verdict: {feedback.validation_verdict.upper()}

Razón:
{feedback.validation_reasoning}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
COMENTARIO DEL TRAINER
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{escalation_reason}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DATOS (primeros registros)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{data_preview}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ID del Feedback: {feedback.id}
"""
    
    # Enviar
    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = admin_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
    
    logger.info(f"✅ Email enviado a {admin_email}")

@app.route('/api/trainer/feedback', methods=['GET'])
@token_required
def get_trainer_feedback(user_id):
    """Obtener feedback guardado con opción de filtrar por categoría"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return safe_jsonify({'error': 'Solo trainers pueden ver feedback'}), 403
        
        # Parámetro opcional para filtrar por categoría
        category = request.args.get('category')
        
        query = session.query(TrainerFeedback).order_by(TrainerFeedback.created_at.desc())
        
        if category:
            query = query.filter_by(category=category)
        
        feedbacks = query.limit(100).all()
        
        result = [{
            "id": f.id,
            "conversation_id": f.conversation_id,
            "original_query": f.original_query,
            "original_response": f.original_response[:100] if f.original_response else None,
            "corrected_response": f.corrected_response[:100] if f.corrected_response else None,
            "feedback_type": f.feedback_type,
            "notes": f.notes,
            "category": f.category,
            "tags": json.loads(f.tags) if f.tags else [],
            "query_type": f.query_type,
            "similarity_score": f.similarity_score,
            "created_at": f.created_at.isoformat()
        } for f in feedbacks]
        
        session.close()
        return safe_jsonify({'success': True, 'feedbacks': result}), 200
        
    except Exception as e:
        return safe_jsonify({'error': str(e)}), 500

@app.route('/api/trainer/feedback/categories', methods=['GET'])
@token_required
def get_feedback_categories(user_id):
    """Obtener lista de categorías disponibles"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return safe_jsonify({'error': 'Solo trainers'}), 403
        
        categories = session.query(TrainerFeedback.category)\
            .filter(TrainerFeedback.category.isnot(None))\
            .distinct()\
            .all()
        
        categories = [c[0] for c in categories]
        session.close()
        
        return safe_jsonify({'success': True, 'categories': categories}), 200
        
    except Exception as e:
        return safe_jsonify({'error': str(e)}), 500

@app.route('/api/trainer/upload', methods=['POST'])
@token_required
def upload_excel(user_id):
    """Subir Excel dinámico - crea tabla automáticamente"""
    session = None
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return safe_jsonify({'error': 'Solo trainers pueden subir archivos'}), 403
        
        if 'file' not in request.files:
            return safe_jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return safe_jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return safe_jsonify({'error': 'Solo archivos .xlsx'}), 400
        
        # Obtener nombre de tabla del archivo
        table_name = file.filename.replace('.xlsx', '').lower()
        table_name = table_name.replace(' ', '_')
        table_name = table_name.replace('-', '_')
        
        # Validar nombre válido
        if not table_name.isidentifier():
            return safe_jsonify({'error': 'Nombre de archivo inválido'}), 400
        
        # Leer Excel
        excel_file = io.BytesIO(file.read())
        df = pd.read_excel(excel_file)
        
        logger.info(f"📊 Columnas del Excel: {list(df.columns)}")
        
        # NUEVO: Verificar si tabla ya existe ANTES de validar 'id'
        table_exists = False
        with engine.connect() as conn:
            stmt = text(f"SELECT 1 FROM information_schema.tables WHERE table_name = '{table_name}'")
            result = conn.execute(stmt).first()
            table_exists = result is not None
        
        # Validar que tenga columna 'id' SOLO si tabla es nueva
        if not table_exists:
            # Tabla nueva: EXIGE 'id'
            if 'id' not in df.columns:
                logger.error(f"❌ Falta columna 'id' en tabla nueva")
                return safe_jsonify({'error': 'Excel debe tener columna "id" para tabla nueva'}), 400
        else:
            # Tabla existe: NO exige 'id'
            logger.info(f"✅ Tabla {table_name} existe, omitiendo validación de 'id'")
        
        # NUEVO: Forzar rollback de transacciones previas
        try:
            with engine.connect() as conn:
                conn.execute(text("ROLLBACK"))
        except:
            pass
        
        # Crear tabla o agregar datos
        try:
            with engine.connect() as conn:
                rows_inserted = 0
                errors = []
                
                if not table_exists:
                    # TABLA NUEVA: CREATE TABLE
                    logger.info(f"📊 Creando tabla nueva: {table_name}")
                    
                    columns_sql = "id SERIAL PRIMARY KEY"
                    
                    for col in df.columns:
                        if col != 'id':
                            if df[col].dtype in ['int64', 'int32']:
                                col_type = 'INTEGER'
                            elif df[col].dtype in ['float64', 'float32']:
                                col_type = 'NUMERIC'
                            else:
                                col_type = 'VARCHAR(255)'
                            
                            columns_sql += f", {col} {col_type}"
                    
                    create_sql = f"CREATE TABLE {table_name} ({columns_sql})"
                    conn.execute(text(create_sql))
                    conn.commit()
                    logger.info(f"✅ Tabla creada: {table_name}")
                else:
                    # TABLA EXISTE: Solo INSERT
                    logger.info(f"📊 Tabla {table_name} existe, agregando datos...")
                
                # Insertar datos (nuevo o existente)
                for idx, row in df.iterrows():
                    try:
                        cols = ", ".join([col for col in df.columns if col != 'id'])
                        placeholders = ", ".join([f"'{str(row[col]).replace(chr(39), chr(39)*2)}'" 
                                                 for col in df.columns if col != 'id'])
                        
                        insert_sql = f"INSERT INTO {table_name} ({cols}) VALUES ({placeholders})"
                        conn.execute(text(insert_sql))
                        rows_inserted += 1
                        log_audit(user_id, user.username, 'UPLOAD', details=f"Tabla: {table_name}, Filas: {rows_inserted}", ip_address=request.remote_addr)
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                
                conn.commit()
        
        except Exception as e:
            logger.error(f"Error en tabla: {e}")
            raise
        
        # Guardar metadata (solo si tabla nueva)
        if not table_exists:
            try:
                meta_session = Session()
                dynamic_table = DynamicTable(
                    table_name=table_name,
                    columns=json.dumps(list(df.columns)),
                    created_by=user_id
                )
                meta_session.add(dynamic_table)
                meta_session.commit()
                meta_session.close()
            except Exception as meta_err:
                logger.error(f"Error guardando metadata: {meta_err}")
        
        session.close()
        
        # Mensaje según si es tabla nueva o existente
        if table_exists:
            logger.info(f"✅ Tabla {table_name}: {rows_inserted} filas agregadas")
            message = f"Se agregaron {rows_inserted} filas a la tabla {table_name}"
        else:
            logger.info(f"✅ Tabla dinámica: {table_name} - {rows_inserted} filas")
            message = f"Tabla {table_name} creada con {rows_inserted} filas"
        
        # ENVIAR EMAIL EN AMBOS CASOS
        send_email_notification(table_name, list(df.columns), user.username, is_new_table=not table_exists)
        
        return safe_jsonify({
            'success': True,
            'table_name': table_name,
            'rows_inserted': rows_inserted,
            'columns': list(df.columns),
            'errors': errors,
            'message': message,
            'table_exists': table_exists
        }), 200
        
    except Exception as e:
        logger.error(f"Error uploadando Excel: {e}")
        if session:
            session.close()
        return safe_jsonify({'error': str(e)}), 500
    

@app.route('/api/export/chart', methods=['POST'])
@token_required
def export_chart(user_id):
    """Exportar gráfico como PNG"""
    try:
        data = request.json
        chart_type = data.get('chart_type')
        chart_data = data.get('chart_data')
        
        if not chart_type or not chart_data:
            return safe_jsonify({'error': 'Datos incompletos'}), 400
        
        
        # Crear gráfico con Plotly
        if chart_type == 'bar':
            labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
            values = [row.get('facturacion', 0) for row in chart_data]
            
            fig = go.Figure(data=[go.Bar(x=labels, y=values, marker_color='#58a6ff')])
            fig.update_layout(
                title='Gráfico de Barras',
                xaxis_title='Clientes',
                yaxis_title='Facturación',
                template='plotly_dark',
                height=500
            )
        
        elif chart_type == 'pie':
            labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
            values = [row.get('market_share', 0) for row in chart_data]
            
            fig = go.Figure(data=[go.Pie(labels=labels, values=values)])
            fig.update_layout(
                title='Gráfico Circular',
                template='plotly_dark',
                height=500
            )
        
        elif chart_type == 'line':
            labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
            values = [row.get('facturacion', 0) for row in chart_data]
            
            fig = go.Figure(data=[go.Scatter(x=labels, y=values, mode='lines+markers', marker_color='#58a6ff')])
            fig.update_layout(
                title='Gráfico de Línea',
                xaxis_title='Clientes',
                yaxis_title='Facturación',
                template='plotly_dark',
                height=500
            )
        
        # Convertir a PNG
        img_bytes = fig.to_image(format="png")
        img_base64 = base64.b64encode(img_bytes).decode()
        
        return safe_jsonify({
            'success': True,
            'image': img_base64,
            'filename': f'grafico_{chart_type}.png'
        }), 200
        
    except Exception as e:
        logger.error(f"Error exportando gráfico: {e}")
        return safe_jsonify({'error': str(e)}), 500

@app.route('/api/export/pdf', methods=['POST'])
@token_required
def export_pdf(user_id):
    """Exportar respuesta + gráfico como PDF"""
    try:
        data = request.json
        response = data.get('response', '')
        chart_type = data.get('chart_type')
        chart_data = data.get('chart_data')
        
        # Crear PDF en memoria
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor='#58a6ff',
            spaceAfter=12
        )
        
        # Título
        elements.append(Paragraph('JARVIS - Reporte', title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Fecha
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        elements.append(Paragraph(f'<b>Fecha:</b> {fecha}', styles['Normal']))
        elements.append(Spacer(1, 0.4*inch))
        
        # Respuesta
        elements.append(Paragraph(response, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Generar gráfico si existe
        if chart_type and chart_data:
            try:
                # Crear gráfico con Plotly
                if chart_type == 'bar':
                    labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
                    values = [row.get('facturacion', 0) for row in chart_data]
                    
                    fig = go.Figure(data=[go.Bar(x=labels, y=values, marker_color='#58a6ff')])
                    fig.update_layout(
                        title='Gráfico de Barras',
                        xaxis_title='Clientes',
                        yaxis_title='Facturación',
                        template='plotly_dark',
                        height=400,
                        width=600
                    )
                
                elif chart_type == 'pie':
                    labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
                    values = [row.get('market_share', 0) for row in chart_data]
                    
                    fig = go.Figure(data=[go.Pie(labels=labels, values=values)])
                    fig.update_layout(
                        title='Gráfico Circular',
                        template='plotly_dark',
                        height=400,
                        width=600
                    )
                
                elif chart_type == 'line':
                    labels = [row.get('cliente', row.get('name', '')) for row in chart_data]
                    values = [row.get('facturacion', 0) for row in chart_data]
                    
                    fig = go.Figure(data=[go.Scatter(x=labels, y=values, mode='lines+markers', marker_color='#58a6ff')])
                    fig.update_layout(
                        title='Gráfico de Línea',
                        xaxis_title='Clientes',
                        yaxis_title='Facturación',
                        template='plotly_dark',
                        height=400,
                        width=600
                    )
                
                # Convertir a PNG
                img_bytes = fig.to_image(format="png")
                img_buffer = BytesIO(img_bytes)
                
                # Agregar imagen al PDF
                elements.append(Spacer(1, 0.2*inch))
                elements.append(RLImage(img_buffer, width=5.5*inch, height=3.3*inch))
                
            except Exception as e:
                logger.error(f"Error generando gráfico en PDF: {e}")
                elements.append(Paragraph(f'<i>Error al incluir gráfico: {str(e)}</i>', styles['Normal']))
        
        # Generar PDF
        doc.build(elements)
        pdf_buffer.seek(0)
        pdf_base64 = base64.b64encode(pdf_buffer.read()).decode()
        
        return safe_jsonify({
            'success': True,
            'pdf': pdf_base64,
            'filename': f'reporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        }), 200
        
    except Exception as e:
        logger.error(f"Error exportando PDF: {e}")
        return safe_jsonify({'error': str(e)}), 500


@app.route('/api/export/excel', methods=['POST'])
@token_required
def export_excel(user_id):
    """Exportar datos como Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        import base64
        from datetime import datetime
        
        data = request.json
        table_data = data.get('data', [])
        filename = data.get('filename', 'datos')
        
        if not table_data:
            return safe_jsonify({'error': 'Sin datos para exportar'}), 400
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Headers
        headers = list(table_data[0].keys()) if table_data else []
        header_fill = PatternFill(start_color='58a6ff', end_color='58a6ff', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, row_data in enumerate(table_data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = row_data.get(header, '')
                cell.alignment = Alignment(horizontal='left')
        
        # Auto-ajustar ancho
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_base64 = base64.b64encode(excel_buffer.read()).decode()
        
        return safe_jsonify({
            'success': True,
            'excel': excel_base64,
            'filename': f'{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        }), 200
        
    except Exception as e:
        logger.error(f"Error exportando Excel: {e}")
        return safe_jsonify({'error': str(e)}), 500
    
@app.route('/api/audit/logs', methods=['GET'])
@token_required
def get_audit_logs(user_id):
    """Obtener logs de auditoría"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return safe_jsonify({'error': 'Solo trainers pueden ver auditoría'}), 403
        
        logs = session.query(AuditLog)\
            .order_by(AuditLog.created_at.desc())\
            .limit(500)\
            .all()
        
        result = [{
            "id": log.id,
            "user_id": log.user_id,
            "username": log.username,
            "action": log.action,
            "details": log.details,
            "ip_address": log.ip_address,
            "created_at": log.created_at.isoformat()
        } for log in logs]
        
        session.close()
        return safe_jsonify({'success': True, 'logs': result}), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo logs: {e}")
        return safe_jsonify({'error': str(e)}), 500

@app.route('/api/trainer/tables', methods=['GET'])
@token_required

def get_tables(user_id):
    """Obtener lista de tablas dinámicas creadas por usuarios"""
    try:
        session = Session()
        user = session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            session.close()
            return safe_jsonify({'error': 'Solo trainers'}), 403
        
        # Obtener solo tablas dinámicas creadas por usuarios (NO del sistema)
        tables_from_db = session.query(DynamicTable)\
            .filter(~DynamicTable.table_name.in_([
                'users', 'conversations', 'trainer_feedback', 'dynamic_tables', 
                'audit_logs', 'dim_anunciante', 'dim_anunciante_alias', 
                'dim_anunciante_perfil', 'fac_facturacion', 'jarvis_query_logs'
            ]))\
            .all()
        
        # Verificar que las tablas existan en la BD (sincronizar)
        valid_tables = []
        for t in tables_from_db:
            try:
                with engine.connect() as conn:
                    stmt = text(f"SELECT 1 FROM information_schema.tables WHERE table_name = '{t.table_name}'")
                    exists = conn.execute(stmt).first()
                    if exists:
                        valid_tables.append(t)
                    else:
                        # Tabla no existe, eliminarla de DynamicTable
                        logger.info(f"🗑️ Tabla {t.table_name} no existe en BD, eliminando de metadata")
                        session.query(DynamicTable).filter_by(table_name=t.table_name).delete()
                        session.commit()
            except Exception as sync_err:
                logger.error(f"Error sincronizando tabla {t.table_name}: {sync_err}")
        
        result = [{
            "table_name": t.table_name,
            "columns": json.loads(t.columns) if t.columns else [],
            "created_at": t.created_at.isoformat() if t.created_at else ""
        } for t in valid_tables]
        
        session.close()
        return safe_jsonify({'success': True, 'tables': result}), 200
    except Exception as e:
        logger.error(f"Error obteniendo tablas: {e}")
        return safe_jsonify({'error': str(e)}), 500


@app.route('/api/trainer/export-template', methods=['POST'])
@token_required
def export_template(user_id):
    """Exportar template de tabla"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        import base64
        
        user_session = Session()
        user = user_session.query(User).filter_by(id=user_id).first()
        
        if not user or user.role != 'trainer':
            user_session.close()
            return safe_jsonify({'error': 'Solo trainers'}), 403
        
        data = request.json
        table_name = data.get('table_name')
        
        if not table_name:
            return safe_jsonify({'error': 'table_name requerido'}), 400
        
        # Obtener columnas de la tabla de PostgreSQL
        with engine.connect() as conn:
            stmt = text(f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = '{table_name}'
                ORDER BY ordinal_position
            """)
            result = conn.execute(stmt)
            columns = [row[0] for row in result]
        
        user_session.close()
        
        # Omitir 'id' del template
        columns = [col for col in columns if col != 'id']
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = table_name
        
        # Headers
        header_fill = PatternFill(start_color='58a6ff', end_color='58a6ff', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-ajustar ancho
        for col in ws.columns:
            max_length = len(str(col[0].value)) if col[0].value else 0
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 15)
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_base64 = base64.b64encode(excel_buffer.read()).decode()
        
        return safe_jsonify({
            'success': True,
            'excel': excel_base64,
            'filename': f'{table_name}_template.xlsx'
        }), 200
        
    except Exception as e:
        logger.error(f"Error exportando template: {e}")
        return safe_jsonify({'error': str(e)}), 500

def send_email_notification(table_name, columns, username, is_new_table=True):
    """Enviar email de notificación de nueva tabla o datos agregados"""
    try:
        sender_email = "danilo.sosa@texo.com.py"
        sender_password = "yfvp aiws uorn ycsr"
        receiver_email = "danilo.sosa@texo.com.py"
        
        # Cambiar título y contenido según si es tabla nueva o INSERT
        if is_new_table:
            subject = f"🔔 Nueva tabla creada: {table_name}"
            title = "Nueva Tabla Creada en JARVIS"
            action = "⚠️ Acción requerida: Integra esta tabla en JARVIS si es necesario."
        else:
            subject = f"📊 Datos agregados a tabla: {table_name}"
            title = "Datos Agregados a Tabla en JARVIS"
            action = "ℹ️ Información: Se han agregado datos a la tabla existente."
        
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = sender_email
        message["To"] = receiver_email
        
        html = f"""\
        <html>
          <body style="font-family: Arial, sans-serif;">
            <h2 style="color: #58a6ff;">{title}</h2>
            <table style="border-collapse: collapse; width: 100%;">
              <tr style="background: #21262d;">
                <td style="padding: 8px; border: 1px solid #30363d;"><strong>Tabla:</strong></td>
                <td style="padding: 8px; border: 1px solid #30363d;">{table_name}</td>
              </tr>
              <tr>
                <td style="padding: 8px; border: 1px solid #30363d;"><strong>Columnas:</strong></td>
                <td style="padding: 8px; border: 1px solid #30363d;">{', '.join(columns)}</td>
              </tr>
              <tr style="background: #21262d;">
                <td style="padding: 8px; border: 1px solid #30363d;"><strong>Usuario:</strong></td>
                <td style="padding: 8px; border: 1px solid #30363d;">{username}</td>
              </tr>
              <tr>
                <td style="padding: 8px; border: 1px solid #30363d;"><strong>Fecha:</strong></td>
                <td style="padding: 8px; border: 1px solid #30363d;">{datetime.now().strftime('%d/%m/%Y %H:%M')}</td>
              </tr>
            </table>
            <hr style="border: none; border-top: 1px solid #30363d; margin: 20px 0;">
            <p><strong>{action}</strong></p>
            <p>Cuando esté lista, notifica al usuario con este mensaje:</p>
            <p style="background: #21262d; padding: 10px; border-radius: 4px;"><code>✅ Tabla '{table_name}' ya está disponible para consultas en JARVIS</code></p>
          </body>
        </html>
        """
        
        part = MIMEText(html, "html")
        message.attach(part)
        
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, receiver_email, message.as_string())
        
        logger.info(f"✅ Email enviado: tabla {table_name}")
        return True
    except Exception as e:
        logger.error(f"❌ Error enviando email: {e}")
        return False
        logger.info(f"📧 Intentando enviar email...")
        logger.info(f"De: {sender_email}")
        logger.info(f"Para: {receiver_email}")
        logger.info(f"Tabla: {table_name}")
    
@app.route('/api/query-direct', methods=['POST'])
def query_direct():
    """Endpoint directo sin Claude - solo BD"""
    try:
        data = request.get_json()
        query = data.get('query', '').lower()
        
        # Solo para rankings
        if not any(word in query for word in ['top', 'ranking', 'clientes', 'principal']):
            return jsonify({"error": "Solo rankings directos"}), 400
        
        with engine.connect() as conn:
            stmt = text("""
                SELECT 
                    cliente_original,
                    COUNT(*) as registros,
                    SUM(facturacion) as facturacion_total
                FROM fact_facturacion
                WHERE facturacion > 0 
                    AND anio = 2025
                    AND cliente_original IS NOT NULL
                GROUP BY cliente_original
                ORDER BY facturacion_total DESC
                LIMIT 10
            """)
            rows = conn.execute(stmt).fetchall()
            
            # Crear respuesta directa HTML
            html_response = """
            <div style="font-family: Arial, sans-serif; margin: 20px;">
            <h2>🏆 Top 10 Clientes por Facturación 2025 (DATOS DIRECTOS BD)</h2>
            <table style="border-collapse: collapse; width: 100%; margin-top: 20px;">
            <tr style="background-color: #f0f0f0; border: 1px solid #ddd;">
                <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">#</th>
                <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">Cliente</th>
                <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Registros</th>
                <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Facturación (Gs)</th>
            </tr>
            """
            
            for i, row in enumerate(rows, 1):
                html_response += f"""
                <tr style="border: 1px solid #ddd;">
                    <td style="padding: 10px; border: 1px solid #ddd;">{i}</td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{row[0]}</td>
                    <td style="padding: 10px; text-align: right; border: 1px solid #ddd;">{row[1]:,}</td>
                    <td style="padding: 10px; text-align: right; border: 1px solid #ddd;">{float(row[2]):,.0f}</td>
                </tr>
                """
            
            html_response += """
            </table>
            <p style="margin-top: 20px; color: #666; font-size: 14px;">
            ✅ Datos obtenidos directamente de la base de datos sin procesamiento de Claude.<br>
            ✅ CERVEPAR debe mostrar: 1,136 registros y 26,057,164,652 Gs<br>
            ✅ TELEFÓNICA debe mostrar: 186 registros y 9,213,656,412 Gs
            </p>
            </div>
            """
            
            return jsonify({
                "success": True,
                "response": html_response,
                "data_source": "BD_DIRECTA"
            })
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/debug-connection', methods=['GET'])
def debug_connection():
    """Debug de conexión"""
    try:
        with engine.connect() as conn:
            # Info de conexión
            db_info = conn.execute(text("SELECT current_database(), current_user")).fetchone()
            
            # Total registros
            total = conn.execute(text("SELECT COUNT(*) FROM fact_facturacion")).fetchone()
            
            # CERVEPAR específico
            cervepar = conn.execute(text("""
                SELECT COUNT(*), SUM(facturacion) 
                FROM fact_facturacion 
                WHERE cliente_original = 'CERVEPAR S.A.'
            """)).fetchone()
            
            return jsonify({
                "database": db_info[0],
                "user": db_info[1],
                "total_registros": total[0],
                "cervepar_registros": cervepar[0],
                "cervepar_facturacion": float(cervepar[1]),
                "connection_string": f"postgresql://postgres:***@localhost/{db_info[0]}"
            })
    except Exception as e:
        return jsonify({"error": str(e)})
    

@app.route('/api/query-direct-fixed', methods=['POST'])
def query_direct_fixed():
    """Endpoint directo CORREGIDO - misma consulta que debug"""
    try:
        data = request.get_json()
        query = data.get('query', '').lower()
        
        if not any(word in query for word in ['top', 'ranking', 'clientes', 'principal']):
            return jsonify({"error": "Solo rankings directos"}), 400
        
        with engine.connect() as conn:
            # LA MISMA CONSULTA QUE USA EL DEBUG QUE FUNCIONA
            stmt = text("""
                SELECT 
                    cliente_original,
                    COUNT(*) as registros,
                    SUM(facturacion) as facturacion_total
                FROM fact_facturacion
                WHERE cliente_original IS NOT NULL
                GROUP BY cliente_original
                ORDER BY facturacion_total DESC
                LIMIT 10
            """)
            rows = conn.execute(stmt).fetchall()
            
            html_response = """
            <div style="font-family: Arial, sans-serif; margin: 20px;">
            <h2>🏆 Top 10 Clientes CORREGIDO (MISMA CONSULTA QUE DEBUG)</h2>
            <table style="border-collapse: collapse; width: 100%; margin-top: 20px;">
            <tr style="background-color: #f0f0f0; border: 1px solid #ddd;">
                <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">#</th>
                <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">Cliente</th>
                <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Registros</th>
                <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Facturación (Gs)</th>
            </tr>
            """
            
            for i, row in enumerate(rows, 1):
                html_response += f"""
                <tr style="border: 1px solid #ddd;">
                    <td style="padding: 10px; border: 1px solid #ddd;">{i}</td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{row[0]}</td>
                    <td style="padding: 10px; text-align: right; border: 1px solid #ddd;">{row[1]:,}</td>
                    <td style="padding: 10px; text-align: right; border: 1px solid #ddd;">{float(row[2]):,.0f}</td>
                </tr>
                """
            
            html_response += """
            </table>
            <p style="margin-top: 20px; color: #666; font-size: 14px;">
            ✅ CONSULTA CORREGIDA - Misma que funciona en debug<br>
            ✅ Debería mostrar CERVEPAR: 1,136 registros, 26,057,164,652 Gs
            </p>
            </div>
            """
            
            return jsonify({
                "success": True,
                "response": html_response,
                "data_source": "BD_CORREGIDA"
            })
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
if __name__ == '__main__':
    logger.info("🚀 JARVIS Backend + Tablas Dinámicas iniciando...")
    app.run(host='0.0.0.0', port=5000, debug=True)